import re
import secrets
import shutil
import string
import time
from datetime import date
from os import path

import openpyxl
from Pass import Pass
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select


class AccessCreator:
    def __init__(self, fullname, title):
        self.fullname = fullname
        fullname = fullname.split()
        self.username = fullname[0] + '.' + fullname[1]
        self.firstName = fullname[0]
        self.lastName = fullname[1]
        self.buro_email = self.username + "@buroserv.com.au"
        self.planettel_email = self.username + "@planettel.com.au"
        self.title = title
        self.logbook = []  # tuples of the form (row, username, password)
        self.logins = Pass()

        # Set up webdriver
        self.driver = webdriver.Chrome(r"Drivers/chromedriver.exe")
        self.driver.maximize_window()
        self.driver.implicitly_wait(10)

    def passwordGenerator(self):
        punctuation = "!@#$%^&*-+?"
        characters = string.ascii_letters + string.digits + punctuation
        password = ""
        while self.passwordCheck(password) is not True:
            password = ""
            for x in range(8):
                password += secrets.choice(characters)
        return password

    def passwordCheck(self, password):
        if re.search(r"[!@#$%^&*\-+?]", password) and re.search(r"[A-Z]", password) and re.search(r"[a-z]",password) and re.search(r"[0-9]", password):
            return True
        return False

    def ims(self):
        self.driver.get(r"http://ims.buroserv.com.au/login.php")

        username = self.driver.find_element_by_id("login-username")
        password = self.driver.find_element_by_id("login-password")
        username.send_keys(self.logins.ims[0])
        password.send_keys(self.logins.ims[1])
        self.driver.find_element_by_id("btn-login").click()

        time.sleep(2)
        self.driver.get(r"http://ims.buroserv.com.au/users.php")
        self.driver.find_element_by_link_text("Add User").click()
        email = self.driver.find_element_by_id("adduser-email")
        username = self.driver.find_element_by_id("adduser-username")
        password = self.driver.find_element_by_id("adduser-password")
        confirm_password = self.driver.find_element_by_id("adduser-confirm_password")
        firstName = self.driver.find_element_by_id("adduser-first_name")
        lastName = self.driver.find_element_by_id("adduser-last_name")
        email.send_keys(self.buro_email)
        username.send_keys(self.username)
        password.send_keys("Temp123")
        confirm_password.send_keys("Temp123")
        firstName.send_keys(self.firstName)
        lastName.send_keys(self.lastName)

        self.driver.find_element_by_id("btn-add-user").click()

        print("Please choose a suitable role for the user.")
        input("Press enter when done and the script will proceed.")

        # Check if user exists after creation
        search = self.driver.find_element_by_xpath('//*[@id="users-list_filter"]/label/input')
        search.send_keys(self.username)
        users_list = self.driver.find_element_by_id("users-list").text
        if self.username in users_list:
            self.logbook.append((17, self.username, "Temp123"))
            print(self.username)
            print("Temp123")
            print("IMS user successfully created...")
        else:
            print("IMS user not created!")

    def tele_iboss(self):
        self.driver.get(r"https://symbio-aspire.iboss.com.au/aspireV2/login.php")
        username = self.driver.find_element_by_name("Username")
        password = self.driver.find_element_by_name("Password")
        username.send_keys(self.logins.tele_iboss[0])
        password.send_keys(self.logins.tele_iboss[1])
        self.driver.find_element_by_name("Submit").click()
        self.driver.find_element_by_id("ToolsDrop").click()
        self.driver.find_element_by_link_text("Wholesaler User Logins").click()

        username = self.driver.find_element_by_id("Username")
        password = self.driver.find_element_by_id("Password")
        name = self.driver.find_element_by_id("Name")
        email = self.driver.find_element_by_id("Email")
        iboss_username = self.firstName + self.lastName
        username.send_keys(iboss_username)
        newPass = self.passwordGenerator()
        special_char = secrets.choice("!#%@<>")
        newPass = re.sub(r"[$%^&*\-+?]", special_char, newPass)
        password.send_keys(newPass)
        email.send_keys(self.buro_email)
        name.send_keys(self.fullname)

        self.driver.find_element_by_name("Submit").click()

        # Check if user exists after creation
        users_list = self.driver.find_element_by_id("WholesellerUserID").text
        if self.fullname in users_list:
            self.logbook.append((31, iboss_username, newPass))
            print(iboss_username)
            print(newPass)
            print("IBoss user successfully created...")
        else:
            print("IBoss user not created!")

    def viaip_optus(self):
        self.driver.get("https://www2.optus.com.au/wholesalenet/")
        username = self.driver.find_element_by_id("USER")
        password = self.driver.find_element_by_id("PASSWORD")
        username.send_keys(self.logins.viaip_optus[0])
        password.send_keys(self.logins.viaip_optus[1])
        self.driver.find_element_by_name("LOGIN").click()
        self.driver.find_element_by_link_text("Administration").click()
        self.driver.find_element_by_xpath('//span[text()="User Management"]').click()

        firstName = self.driver.find_element_by_id("FirstName")
        lastName = self.driver.find_element_by_id("LastName")
        username = self.driver.find_element_by_id("Username")
        email = self.driver.find_element_by_name("Email")
        firstName.send_keys(self.firstName)
        lastName.send_keys(self.lastName)
        username.send_keys(".VIAIP")
        email.send_keys(self.buro_email)
        # js drop-down boxes
        self.driver.find_element_by_xpath('//*[@id="addNewUser"]/ul[1]/li[1]/span/button').click()
        self.driver.find_element_by_xpath('/html/body/div[4]/ul/li[2]/label').click()
        self.driver.find_element_by_xpath(
            '//*[@id="addNewUser"]/ul[1]/li[6]/span/table[1]/tbody/tr[3]/td[2]/button').click()
        self.driver.find_element_by_xpath('/html/body/div[6]/ul/li[2]/label').click()
        self.driver.find_element_by_xpath(
            '//*[@id="addNewUser"]/ul[1]/li[6]/span/table[1]/tbody/tr[4]/td[2]/button').click()
        self.driver.find_element_by_xpath('/html/body/div[7]/ul/li[2]/label').click()
        self.driver.find_element_by_xpath(
            '//*[@id="addNewUser"]/ul[1]/li[6]/span/table[1]/tbody/tr[5]/td[2]/button').click()
        self.driver.find_element_by_xpath('/html/body/div[8]/ul/li[2]/label').click()
        self.driver.find_element_by_xpath(
            '//*[@id="addNewUser"]/ul[1]/li[6]/span/table[1]/tbody/tr[6]/td[2]/button').click()
        self.driver.find_element_by_xpath('/html/body/div[9]/ul/li[2]/label').click()
        self.driver.find_element_by_xpath(
            '//*[@id="addNewUser"]/ul[1]/li[6]/span/table[1]/tbody/tr[7]/td[2]/button').click()
        self.driver.find_element_by_xpath('/html/body/div[10]/ul/li[2]/label').click()
        self.driver.find_element_by_xpath(
            '//*[@id="addNewUser"]/ul[1]/li[6]/span/table[1]/tbody/tr[8]/td[2]/button').click()
        self.driver.find_element_by_xpath('/html/body/div[11]/ul/li[2]/label').click()
        self.driver.find_element_by_xpath(
            '//*[@id="addNewUser"]/ul[1]/li[6]/span/table[1]/tbody/tr[9]/td[2]/button').click()
        self.driver.find_element_by_xpath('/html/body/div[12]/ul/li[2]/label').click()
        self.driver.find_element_by_xpath(
            '//*[@id="addNewUser"]/ul[1]/li[6]/span/table[1]/tbody/tr[10]/td[2]/button').click()
        self.driver.find_element_by_xpath('/html/body/div[13]/ul/li[2]/label').click()

        # Submit
        self.driver.find_element_by_xpath('//*[@id="addNewUser"]/ul[2]/li[4]/input[1]').click()

        # Check if user exists after creation
        self.driver.find_element_by_link_text("Administration").click()
        element1 = self.driver.find_element_by_xpath('//span[text()="User Management"]')
        element2 = self.driver.find_element_by_xpath('//*[@id="sec_1870"]/li/div/div/div/ul/li[2]/a')
        hover = ActionChains(self.driver).move_to_element(element1).move_to_element(element2).click()
        hover.perform()
        self.driver.find_element_by_xpath('//*[@id="content"]/div/div/div[1]/div[2]/a/h2').click()
        username = self.driver.find_element_by_id("Username")
        username.send_keys(self.username + '.VIAIP')
        username.submit()
        try:
            users_list = self.driver.find_element_by_xpath('//*[@id="content"]/div/div/div[1]/div[7]/table').text
            if self.buro_email in users_list:
                self.logbook.append((30, self.username + '.VIAIP', None))
                print("ViaIP Optus user successfully created...")
        except:
            print("ViaIP Optus user not created!")

    # NOT UPDATED DO NOT USE...
    def buro_optus(self):
        self.driver.get("https://www2.optus.com.au/wholesalenet/")
        username = self.driver.find_element_by_id("USER")
        password = self.driver.find_element_by_id("PASSWORD")
        username.send_keys(self.logins.buro_optus[0])
        password.send_keys(self.logins.buro_optus[1])
        self.driver.find_element_by_name("LOGIN").click()
        self.driver.find_element_by_link_text("Administration").click()
        self.driver.find_element_by_xpath('//span[text()="User Management"]').click()

        firstName = self.driver.find_element_by_id("FirstName")
        lastName = self.driver.find_element_by_id("LastName")
        username = self.driver.find_element_by_id("Username")
        email = self.driver.find_element_by_name("Email")
        firstName.send_keys(self.firstName)
        lastName.send_keys(self.lastName)
        username.send_keys(".BURO")
        email.send_keys(self.buro_email)
        # js drop-down boxes
        self.driver.find_element_by_xpath('//*[@id="addNewUser"]/ul[1]/li[1]/span/button').click()
        self.driver.find_element_by_xpath('/html/body/div[4]/ul/li[2]/label').click()
        self.driver.find_element_by_xpath(
            '//*[@id="addNewUser"]/ul[1]/li[6]/span/table[1]/tbody/tr[3]/td[2]/button').click()
        self.driver.find_element_by_xpath('/html/body/div[6]/ul/li[2]/label').click()
        self.driver.find_element_by_xpath(
            '//*[@id="addNewUser"]/ul[1]/li[6]/span/table[1]/tbody/tr[4]/td[2]/button').click()
        self.driver.find_element_by_xpath('/html/body/div[7]/ul/li[2]/label').click()
        self.driver.find_element_by_xpath(
            '//*[@id="addNewUser"]/ul[1]/li[6]/span/table[1]/tbody/tr[5]/td[2]/button').click()
        self.driver.find_element_by_xpath('/html/body/div[8]/ul/li[2]/label').click()
        self.driver.find_element_by_xpath(
            '//*[@id="addNewUser"]/ul[1]/li[6]/span/table[1]/tbody/tr[6]/td[2]/button').click()
        self.driver.find_element_by_xpath('/html/body/div[9]/ul/li[2]/label').click()
        self.driver.find_element_by_xpath(
            '//*[@id="addNewUser"]/ul[1]/li[6]/span/table[1]/tbody/tr[7]/td[2]/button').click()
        self.driver.find_element_by_xpath('/html/body/div[10]/ul/li[2]/label').click()

        # Submit
        self.driver.find_element_by_xpath('//*[@id="addNewUser"]/ul[2]/li[4]/input[1]').click()

    def octane(self):
        self.driver.get(r"https://octane.telcoinabox.com/tiab/Login")
        self.driver.find_element_by_id("login_button").click()
        time.sleep(2)
        username = self.driver.find_element_by_id("j_username")
        password = self.driver.find_element_by_id("predigpass")
        username.send_keys(self.logins.octane[0])
        password.send_keys(self.logins.octane[1])
        password.send_keys(Keys.ENTER)
        self.driver.find_element_by_xpath('//span[text()="Login"]').click()
        self.driver.get(r"https://octane.telcoinabox.com/tiab/NewUser.jsp")
        username = self.driver.find_element_by_id("uid")
        password = self.driver.find_element_by_id("predigpass")
        username.send_keys(self.username)
        newPass = self.passwordGenerator()
        password.send_keys(newPass)

        self.driver.find_element_by_id("submit-button").click()

        # Check if user exists after creation
        self.driver.get(r"https://octane.telcoinabox.com/tiab/UserList")
        users_list = self.driver.find_element_by_xpath('//*[@id="user-list"]/div[2]/table/tbody').text
        if self.username in users_list:
            self.logbook.append((7, self.username, newPass))
            print(self.username)
            print(newPass)
            print("Octane user successfully created...")
        else:
            print("Octane user not created!")
        # TODO: add partitions??

    def clarus_genex(self):
        self.driver.get(r"https://genex.billing.com.au/Module/Main/login.aspx")
        database = self.driver.find_element_by_id("ctl00_CPH_txtDatabase")
        username = self.driver.find_element_by_id("ctl00_CPH_txtUsername")
        password = self.driver.find_element_by_id("ctl00_CPH_txtPassword")
        database.clear()
        username.clear()
        database.send_keys("Clarus")
        username.send_keys(self.logins.clarus_genex[0])
        password.send_keys(self.logins.clarus_genex[1])
        self.driver.find_element_by_id("ctl00_CPH_btnLogin").click()
        self.driver.get(r"https://genex.billing.com.au/module/Roles/UserManager.aspx")
        self.driver.find_element_by_id("ctl00_CPH_UsersAndRoles_AddButton").click()
        username = self.driver.find_element_by_id("ctl00_CPH_UsersAndRoles_txtUserName")
        firstName = self.driver.find_element_by_id("ctl00_CPH_UsersAndRoles_txtFirstName")
        lastName = self.driver.find_element_by_id("ctl00_CPH_UsersAndRoles_txtLastName")
        password = self.driver.find_element_by_id("ctl00_CPH_UsersAndRoles_txtPassword")
        confirm_password = self.driver.find_element_by_id("ctl00_CPH_UsersAndRoles_txtConfirmPassword")
        email = self.driver.find_element_by_id("ctl00_CPH_UsersAndRoles_txtEmailAddress")
        if len(self.username) > 12:
            genex_username = self.firstName + self.lastName[0]
            if len(genex_username) > 12:
                genex_username = input("Please enter a 12 character username for Genex.")
        else:
            genex_username = self.username
        username.send_keys(genex_username)
        firstName.send_keys(self.firstName)
        lastName.send_keys(self.lastName)
        newPass = self.passwordGenerator()
        password.send_keys(newPass)
        confirm_password.send_keys(newPass)
        email.send_keys(self.buro_email)

        self.driver.find_element_by_id("ctl00_CPH_UsersAndRoles_cblGroups_0").click()

        self.driver.find_element_by_id("ctl00_CPH_UsersAndRoles_cblRoles_13").click()
        self.driver.find_element_by_id("ctl00_CPH_UsersAndRoles_cblRoles_16").click()

        self.driver.find_element_by_id("ctl00_CPH_UsersAndRoles_UpdateButtons_SaveButton").click()
        self.driver.switch_to.alert.accept()

        # Check if user exists after creation
        self.driver.get(r"https://genex.billing.com.au/module/Roles/UserManager.aspx")
        users_list = self.driver.find_element_by_id("ctl00_CPH_UsersAndRoles_rblUsers").text
        if self.fullname in users_list:
            self.logbook.append((8, genex_username, newPass))
            print(genex_username)
            print(newPass)
            print("Clarus Genex user successfully created...")
        else:
            print("Clarus Genex user not created!")

    def buro_genex(self):
        self.driver.get(r"https://genex.billing.com.au/Module/Main/login.aspx")
        database = self.driver.find_element_by_id("ctl00_CPH_txtDatabase")
        username = self.driver.find_element_by_id("ctl00_CPH_txtUsername")
        password = self.driver.find_element_by_id("ctl00_CPH_txtPassword")
        database.clear()
        username.clear()
        database.send_keys("Buroserv")
        username.send_keys(self.logins.buro_genex[0])
        password.send_keys(self.logins.buro_genex[1])
        self.driver.find_element_by_id("ctl00_CPH_btnLogin").click()
        self.driver.get(r"https://genex.billing.com.au/module/Roles/UserManager.aspx")
        self.driver.find_element_by_id("ctl00_CPH_UsersAndRoles_AddButton").click()
        username = self.driver.find_element_by_id("ctl00_CPH_UsersAndRoles_txtUserName")
        firstName = self.driver.find_element_by_id("ctl00_CPH_UsersAndRoles_txtFirstName")
        lastName = self.driver.find_element_by_id("ctl00_CPH_UsersAndRoles_txtLastName")
        password = self.driver.find_element_by_id("ctl00_CPH_UsersAndRoles_txtPassword")
        confirm_password = self.driver.find_element_by_id("ctl00_CPH_UsersAndRoles_txtConfirmPassword")
        email = self.driver.find_element_by_id("ctl00_CPH_UsersAndRoles_txtEmailAddress")
        if len(self.username) > 12:
            genex_username = self.firstName + self.lastName[0]
            if len(genex_username) > 12:
                genex_username = input("Please enter a 12 character username for Genex.")
        else:
            genex_username = self.username
        username.send_keys(genex_username)
        firstName.send_keys(self.firstName)
        lastName.send_keys(self.lastName)
        newPass = self.passwordGenerator()
        password.send_keys(newPass)
        confirm_password.send_keys(newPass)
        email.send_keys(self.buro_email)
        self.driver.find_element_by_id("ctl00_CPH_UsersAndRoles_cblGroups_0").click()

        self.driver.find_element_by_id("ctl00_CPH_UsersAndRoles_cblRoles_2").click()
        self.driver.find_element_by_id("ctl00_CPH_UsersAndRoles_cblRoles_5").click()
        self.driver.find_element_by_id("ctl00_CPH_UsersAndRoles_cblRoles_6").click()
        self.driver.find_element_by_id("ctl00_CPH_UsersAndRoles_cblRoles_7").click()
        self.driver.find_element_by_id("ctl00_CPH_UsersAndRoles_cblRoles_8").click()
        self.driver.find_element_by_id("ctl00_CPH_UsersAndRoles_cblRoles_11").click()
        self.driver.find_element_by_id("ctl00_CPH_UsersAndRoles_cblRoles_12").click()
        self.driver.find_element_by_id("ctl00_CPH_UsersAndRoles_cblRoles_16").click()
        self.driver.find_element_by_id("ctl00_CPH_UsersAndRoles_cblRoles_17").click()
        self.driver.find_element_by_id("ctl00_CPH_UsersAndRoles_cblRoles_18").click()

        self.driver.find_element_by_id("ctl00_CPH_UsersAndRoles_UpdateButtons_SaveButton").click()
        self.driver.switch_to.alert.accept()

        # Check if user exists after creation
        self.driver.get(r"https://genex.billing.com.au/module/Roles/UserManager.aspx")
        users_list = self.driver.find_element_by_id("ctl00_CPH_UsersAndRoles_rblUsers").text
        if self.fullname in users_list:
            self.logbook.append((9, genex_username, newPass))
            print(genex_username)
            print(newPass)
            print("Buro Genex user successfully created...")
        else:
            print("Buro Genex user not created!")

    def v4_genex(self):
        self.driver.get(r"https://genex.billing.com.au/Module/Main/login.aspx")
        database = self.driver.find_element_by_id("ctl00_CPH_txtDatabase")
        username = self.driver.find_element_by_id("ctl00_CPH_txtUsername")
        password = self.driver.find_element_by_id("ctl00_CPH_txtPassword")
        database.clear()
        username.clear()
        database.send_keys("v4")
        username.send_keys(self.logins.v4_genex[0])
        password.send_keys(self.logins.v4_genex[1])
        self.driver.find_element_by_id("ctl00_CPH_btnLogin").click()
        self.driver.get(r"https://genex.billing.com.au/module/Roles/UserManager.aspx")
        self.driver.find_element_by_id("ctl00_CPH_UsersAndRoles_AddButton").click()
        username = self.driver.find_element_by_id("ctl00_CPH_UsersAndRoles_txtUserName")
        firstName = self.driver.find_element_by_id("ctl00_CPH_UsersAndRoles_txtFirstName")
        lastName = self.driver.find_element_by_id("ctl00_CPH_UsersAndRoles_txtLastName")
        password = self.driver.find_element_by_id("ctl00_CPH_UsersAndRoles_txtPassword")
        confirm_password = self.driver.find_element_by_id("ctl00_CPH_UsersAndRoles_txtConfirmPassword")
        email = self.driver.find_element_by_id("ctl00_CPH_UsersAndRoles_txtEmailAddress")
        if len(self.username) > 12:
            genex_username = self.firstName + self.lastName[0]
            if len(genex_username) > 12:
                genex_username = input("Please enter a 12 character username for Genex.")
        else:
            genex_username = self.username
        username.send_keys(genex_username)
        firstName.send_keys(self.firstName)
        lastName.send_keys(self.lastName)
        newPass = self.passwordGenerator()
        password.send_keys(newPass)
        confirm_password.send_keys(newPass)
        email.send_keys(self.buro_email)
        self.driver.find_element_by_id("ctl00_CPH_UsersAndRoles_cblGroups_0").click()

        self.driver.find_element_by_id("ctl00_CPH_UsersAndRoles_cblRoles_2").click()
        self.driver.find_element_by_id("ctl00_CPH_UsersAndRoles_cblRoles_5").click()
        self.driver.find_element_by_id("ctl00_CPH_UsersAndRoles_cblRoles_6").click()
        self.driver.find_element_by_id("ctl00_CPH_UsersAndRoles_cblRoles_7").click()
        self.driver.find_element_by_id("ctl00_CPH_UsersAndRoles_cblRoles_8").click()
        self.driver.find_element_by_id("ctl00_CPH_UsersAndRoles_cblRoles_11").click()
        self.driver.find_element_by_id("ctl00_CPH_UsersAndRoles_cblRoles_12").click()
        self.driver.find_element_by_id("ctl00_CPH_UsersAndRoles_cblRoles_16").click()
        self.driver.find_element_by_id("ctl00_CPH_UsersAndRoles_cblRoles_17").click()
        self.driver.find_element_by_id("ctl00_CPH_UsersAndRoles_cblRoles_18").click()

        self.driver.find_element_by_id("ctl00_CPH_UsersAndRoles_UpdateButtons_SaveButton").click()
        self.driver.switch_to.alert.accept()

        # Check if user exists after creation
        self.driver.get(r"https://genex.billing.com.au/module/Roles/UserManager.aspx")
        users_list = self.driver.find_element_by_id("ctl00_CPH_UsersAndRoles_rblUsers").text
        if self.fullname in users_list:
            self.logbook.append((8, genex_username, newPass))
            print(genex_username)
            print(newPass)
            print("V4 Genex user successfully created...")
        else:
            print("V4 Genex user not created!")

    def sonar(self):
        self.driver.get(r"https://mvp02.symbionetworks.com/sonar_admin/")
        username = self.driver.find_element_by_name("j_username")
        password = self.driver.find_element_by_name("j_password")
        username.send_keys(self.logins.sonar[0])
        password.send_keys(self.logins.sonar[1])
        password.send_keys(Keys.ENTER)

        self.driver.find_element_by_class_name("rootVoices").click()
        self.driver.find_element_by_link_text("Users").click()
        self.driver.find_element_by_xpath('//button[text()="Create User"]').click()
        time.sleep(2)  # wait for js to load
        self.driver.find_element_by_id("EndUserActor2671createUserenabled").click()
        username = self.driver.find_element_by_name("username")
        password = self.driver.find_element_by_name("password")
        surname = self.driver.find_element_by_name("surname")
        firstName = self.driver.find_element_by_name("firstName")
        email = self.driver.find_element_by_name("emailAddress")
        dob = self.driver.find_element_by_name("dateOfBirth")
        phone = self.driver.find_element_by_name("phoneNumberBH")
        sonar_username = self.firstName + self.lastName[0]
        username.send_keys(sonar_username)
        newPass = self.passwordGenerator()
        password.send_keys(newPass)
        surname.send_keys(self.lastName)
        firstName.send_keys(self.firstName)
        email.send_keys(self.buro_email)
        dob.send_keys("2019-01-01 00:00")
        phone.send_keys("1")
        self.driver.find_element_by_xpath('//button[text()="Submit"]').click()

        # Check if user exists after creation
        users_list = self.driver.find_element_by_id("EndUserActor2671UsersListingTable").text
        if self.buro_email in users_list:
            self.logbook.append((29, sonar_username, newPass))
            print("Sonar user successfully created...")
        else:
            self.logbook.append((29, sonar_username, newPass))
            print(sonar_username)
            print(newPass)
            print("Sonar user not created!")

    def supatools(self):
        self.driver.get(r"https://support.viptelecombilling.net.au/login.php")
        username = self.driver.find_element_by_name("user_id")
        password = self.driver.find_element_by_name("password")
        username.send_keys(self.logins.supatools[0])
        password.send_keys(self.logins.supatools[1])
        password.send_keys(Keys.ENTER)

        time.sleep(2)  # wait for page to load
        frame = self.driver.find_element_by_xpath('/html/frameset/frameset/frame[1]')
        self.driver.switch_to.frame(frame)
        self.driver.find_element_by_xpath('/html/body/table/tbody/tr[2]/td/a/img').click()
        self.driver.switch_to.default_content()
        frame = self.driver.find_element_by_xpath('/html/frameset/frameset/frame[2]')
        self.driver.switch_to.frame(frame)
        self.driver.find_element_by_xpath(
            '/html/body/table/tbody/tr/td/table/tbody/tr[1]/td/table[2]/tbody/tr/td[4]/a/img').click()
        self.driver.find_element_by_name("newci").click()
        drop_down = Select(
            self.driver.find_element_by_xpath('/html/body/form/table/tbody/tr/td/table/tbody/tr[1]/td[2]/select'))
        drop_down.select_by_visible_text("PlanetTel")
        time.sleep(2)  # wait for page to load
        fullName = self.driver.find_element_by_id("__name")
        title = self.driver.find_element_by_id("__job_title")
        email = self.driver.find_element_by_id("__email")
        self.driver.find_element_by_name("__can_switch_co").click()
        access_role = Select(self.driver.find_element_by_id("__access_role_id"))
        access_role.select_by_visible_text("Planet Tel")
        access_level = Select(self.driver.find_element_by_id("__access_level"))
        access_level.select_by_visible_text("Service Desk")
        username = self.driver.find_element_by_id("__user_id")
        password = self.driver.find_element_by_id("f_password")
        confirm_password = self.driver.find_element_by_id("confirm_password")
        status = Select(self.driver.find_element_by_name("__status"))
        status.select_by_visible_text("Reset")
        send_email = Select(self.driver.find_element_by_id("user_email_template"))
        send_email.select_by_visible_text("user_creation_password.txt")
        fullName.send_keys(self.fullname)
        title.send_keys(self.title)
        email.send_keys(self.buro_email)
        username.send_keys(self.username)
        password.send_keys("Temp123")
        confirm_password.send_keys("Temp123")

        self.driver.find_element_by_id("savebtn").click()

        # Check if user exists after creation
        time.sleep(2)
        self.driver.switch_to.default_content()
        frame = self.driver.find_element_by_xpath('/html/frameset/frame')
        self.driver.switch_to.frame(frame)
        search = self.driver.find_element_by_xpath('//*[@id="filter"]')
        search.send_keys(self.fullname)
        search.send_keys(Keys.ENTER)
        try:
            self.driver.switch_to.default_content()
            frame = self.driver.find_element_by_xpath('/html/frameset/frameset/frame[2]')
            self.driver.switch_to.frame(frame)
            users_list = self.driver.find_element_by_xpath('/html/body/table/tbody/tr/td').text
            if self.username in users_list:
                self.logbook.append((34, self.username, "Temp123"))
                print(self.username)
                print("Temp123")
                print("Supatools user successfully created...")
        except:
            print("Supatools user not created!")

    def porta(self):
        self.driver.get(r"https://billing.isphone.com.au/index.html")
        username = self.driver.find_element_by_id("pb_auth_user")
        password = self.driver.find_element_by_id("pb_auth_password")
        username.send_keys(self.logins.porta[0])
        password.send_keys(self.logins.porta[1])
        password.send_keys(Keys.ENTER)

        self.driver.get(r"https://billing.isphone.com.au/users.html")
        self.driver.find_element_by_class_name("add").click()
        time.sleep(3)  # wait for js to load
        firstName = self.driver.find_element_by_name("firstname")
        lastName = self.driver.find_element_by_id("lastname_field_id")
        email = self.driver.find_element_by_id("email_field_id")
        firstName.send_keys(self.firstName)
        lastName.send_keys(self.lastName)
        email.send_keys(self.buro_email)
        self.driver.find_element_by_link_text("Web Self-Care").click()
        username = self.driver.find_element_by_id("login_input_id")
        password = self.driver.find_element_by_id("password_input_id")
        username.clear()
        username.send_keys(self.username)
        newPass = self.passwordGenerator()
        password.send_keys(newPass)
        drop_down = Select(self.driver.find_element_by_id("ACLSelect"))
        drop_down.select_by_visible_text("ISPhone - Account Manager")
        self.driver.find_element_by_id("api-token-auth").click()

        self.driver.find_element_by_class_name("save_close").click()

        self.driver.get(r"https://billing.isphone.com.au/users.html")
        print("Please verify if user is created (Y/N).")
        answer = input("")
        if answer == 'Y' or answer == 'y':
            self.logbook.append((19, self.username, newPass))
            print(self.username)
            print(newPass)
            print("Porta user successfully created...")
        else:
            print("Porta user not created!")

    def buro_frontier(self):
        self.driver.get(r"https://frontier.aapt.com.au/s/login")
        username = self.driver.find_element_by_name("j_username")
        password = self.driver.find_element_by_name("j_password")
        username.send_keys(self.logins.buro_frontier[0])
        password.send_keys(self.logins.buro_frontier[1])
        self.driver.find_element_by_id("submit").click()

        self.driver.get(r"https://frontier.aapt.com.au/s/manageusers/createUserSelectPerson")
        self.driver.find_element_by_id("createNewUser").click()
        username = self.driver.find_element_by_id("usernameField")
        firstName = self.driver.find_element_by_id("firstNameField")
        lastName = self.driver.find_element_by_id("lastNameField")
        workPhone = self.driver.find_element_by_id("workPhoneField")
        jobTitle = self.driver.find_element_by_id("jobTitleField")
        username.send_keys(self.buro_email)
        firstName.send_keys(self.firstName)
        lastName.send_keys(self.lastName)
        workPhone.send_keys("1300 726 210")
        jobTitle.send_keys(self.title)
        billingRole = self.driver.find_element_by_id("billingRole")
        orderingRole = self.driver.find_element_by_id("orderingRole")
        qualificationRole = self.driver.find_element_by_id("qualificationRole")
        quotingRole = self.driver.find_element_by_id("quotingRole")
        if billingRole.is_selected():
            billingRole.click()  # need to deselect
        if not orderingRole.is_selected():
            orderingRole.click()
        if not qualificationRole.is_selected():
            qualificationRole.click()
        if not quotingRole.is_selected():
            quotingRole.click()

        self.driver.find_element_by_id("createUser").click()

        # Check if user exists after creation
        self.driver.get(r"https://frontier.aapt.com.au/s/manageusers")
        users_list = self.driver.find_element_by_id("userList").text
        match = re.search(self.lastName + ', ' + self.firstName, users_list)
        if match:
            self.logbook.append((23, self.buro_email, None))
            print("Buroserv Frontier user successfully created...")
        else:
            print("Buroserv Frontier user not created!")

        # Sign out
        self.driver.find_element_by_id("nameMenu").click()
        self.driver.find_element_by_link_text("Sign Out").click()

    def cloud_frontier(self):
        self.driver.get(r"https://frontier.aapt.com.au/s/login")
        username = self.driver.find_element_by_name("j_username")
        password = self.driver.find_element_by_name("j_password")
        username.clear()
        username.send_keys(self.logins.cloud_frontier[0])
        password.send_keys(self.logins.cloud_frontier[1])
        self.driver.find_element_by_id("submit").click()

        self.driver.get(r"https://frontier.aapt.com.au/s/manageusers/createUserSelectPerson")
        username = self.driver.find_element_by_id("usernameField")
        firstName = self.driver.find_element_by_id("firstNameField")
        lastName = self.driver.find_element_by_id("lastNameField")
        workPhone = self.driver.find_element_by_id("workPhoneField")
        jobTitle = self.driver.find_element_by_id("jobTitleField")
        username.send_keys(self.planettel_email)
        firstName.send_keys(self.firstName)
        lastName.send_keys(self.lastName)
        workPhone.send_keys("1300 726 210")
        jobTitle.send_keys(self.title)
        billingRole = self.driver.find_element_by_id("billingRole")
        orderingRole = self.driver.find_element_by_id("orderingRole")
        qualificationRole = self.driver.find_element_by_id("qualificationRole")
        quotingRole = self.driver.find_element_by_id("quotingRole")
        if billingRole.is_selected():
            billingRole.click()  # need to deselect
        if not orderingRole.is_selected():
            orderingRole.click()
        if not qualificationRole.is_selected():
            qualificationRole.click()
        if not quotingRole.is_selected():
            quotingRole.click()

        self.driver.find_element_by_id("createUser").click()

        # Check if user exists after creation
        self.driver.get(r"https://frontier.aapt.com.au/s/manageusers")
        users_list = self.driver.find_element_by_id("userList").text
        match = re.search(self.lastName + ', ' + self.firstName, users_list)
        if match:
            self.logbook.append((24, self.planettel_email, None))
            print("Cloudnyne Frontier user successfully created...")
        else:
            print("Cloudnyne Frontier user not created!")

        # Sign out
        self.driver.find_element_by_id("nameMenu").click()
        self.driver.find_element_by_link_text("Sign Out").click()

    def viaip_utilibill(self):
        self.driver.get(r"https://viaip.utilibill.com.au/viaip/Login")
        self.driver.find_element_by_tag_name("a").click()
        username = self.driver.find_element_by_name("j_username")
        password = self.driver.find_element_by_id("predigpass")
        username.send_keys(self.logins.viaip_utilibill[0])
        password.send_keys(self.logins.viaip_utilibill[1])
        self.driver.find_element_by_name("submit").click()
        self.driver.find_element_by_id("submitrequest").click()

        self.driver.get(r"https://viaip.utilibill.com.au/viaip/NewUser.jsp")
        username = self.driver.find_element_by_name("j_username")
        password = self.driver.find_element_by_id("predigpass")
        drop_down = Select(self.driver.find_element_by_id("group"))
        drop_down.select_by_visible_text("ViaIP Administrator")
        username.send_keys(self.username)
        newPass = self.passwordGenerator()
        password.send_keys(newPass)

        self.driver.find_element_by_class_name("buttonLrg").click()

        self.driver.get(r"https://viaip.utilibill.com.au/viaip/UserList")
        users_list = self.driver.find_element_by_id("utbListDiv").text
        match = re.search(self.username, users_list)
        if match:
            self.logbook.append((5, self.username, newPass))
            print(self.username)
            print(newPass)
            print("ViaIP Utilibill user successfully created...")
        else:
            print("ViaIP Utilibill user not created!")

    def cloud_ultilibill(self):
        self.driver.get(r"https://eziconnect.utilibill.com.au/eziconnect/Login")
        self.driver.find_element_by_tag_name("a").click()
        username = self.driver.find_element_by_name("j_username")
        password = self.driver.find_element_by_id("predigpass")
        username.send_keys(self.logins.cloud_utilibill[0])
        password.send_keys(self.logins.cloud_utilibill[1])
        self.driver.find_element_by_name("submit").click()
        self.driver.find_element_by_id("submitrequest").click()

        self.driver.get(r"https://eziconnect.utilibill.com.au/eziconnect/NewUser.jsp")
        username = self.driver.find_element_by_name("j_username")
        password = self.driver.find_element_by_id("predigpass")
        drop_down = Select(self.driver.find_element_by_id("group"))
        drop_down.select_by_visible_text("Eziconnect Administrator")
        username.send_keys(self.username)
        newPass = self.passwordGenerator()
        password.send_keys(newPass)

        self.driver.find_element_by_class_name("buttonLrg").click()

        self.driver.get(r"https://eziconnect.utilibill.com.au/eziconnect/UserList")
        users_list = self.driver.find_element_by_id("utbListDiv").text
        match = re.search(self.username, users_list)
        if match:
            self.logbook.append((6, self.username, newPass))
            print(self.username)
            print(newPass)
            print("Cloudnyne Utilibill user successfully created...")
        else:
            print("Cloudnyne Utilibill user not created!")

    def selcomm(self):
        self.driver.get("https://support.selcomm.com:8443/secure/Dashboard.jspa")
        username = self.driver.find_element_by_id("login-form-username")
        password = self.driver.find_element_by_id("login-form-password")
        username.send_keys(self.logins.selcomm[0])
        password.send_keys(self.logins.selcomm[1])
        self.driver.find_element_by_id("login").click()
        self.driver.find_element_by_id("create_link").click()

        drop_down = Select(self.driver.find_element_by_id("customfield_11502"))
        drop_down.select_by_visible_text("Bryan Yeung")
        summary = self.driver.find_element_by_id("summary")
        description = self.driver.find_element_by_id("description")
        summary.send_keys("New access required")
        description.send_keys("Hi Team,\n\nCould you please create an access to BU & BW for:\n" +
                              self.fullname + " (" + self.buro_email + ")\n" + "Please mirror the access levels of: "
                              + "\n\nThanks!")
        print("Please name a person's access levels to mirror, press Enter when done.")
        input("")
        self.driver.find_element_by_id("create-issue-submit").click()

    def export(self):
        src = path.realpath("New Employee Access Template.xlsx")
        new_wb = self.fullname + " Accesses " + str(date.today()) + ".xlsx"
        shutil.copy(src, new_wb)
        wb = openpyxl.load_workbook(new_wb)
        ws = wb.active
        ws.title = self.fullname + " (" + self.title + ")"

        # Logbook = (row, username, password)
        for x in self.logbook:
            username_row = 'C' + str(x[0])
            password_row = 'D' + str(x[0])
            ws[username_row] = x[1]
            if x[2] is not None:
                ws[password_row] = x[2]
        ws['C20'] = self.buro_email
        ws['D20'] = '1qazXSW@'
        ws['C21'] = self.planettel_email
        wb.save(new_wb)
        shutil.copy(new_wb, r"C:\Users\Bryan\OneDrive - Buroserv Australia Pty Ltd\Documents\New Employee Accesses")

    def createAll(self):
        try:
            # self.ims()
            self.viaip_optus()
            # self.buro_optus()
            self.supatools()
            self.buro_frontier()
            # self.cloud_frontier()
            self.viaip_utilibill()
            self.cloud_ultilibill()
            # self.clarus_genex()
            self.buro_genex()
            self.v4_genex()
            self.octane()
            self.sonar()
            self.tele_iboss()
            self.porta()
            # self.selcomm()

        except:
            print("An Error has occurred! Aborting and writing to excel sheet.")
            print("Please manually complete the rest of the accesses...")
            self.export()
            return

        print("Writing user credentials to a spreadsheet...")
        self.export()

    def teardown(self):
        self.driver.quit()


def main():
    print("Enter the full name of the new user:")
    name = input("")
    print("Enter the job title of the user:")
    title = input("")
    print("Is the following correct? (Y/N)")
    print(name + ",", title)
    answer = input("")
    if answer == 'N' or answer == 'n':
        print("Please run the script again.")
        return

    elif answer == "Y" or answer == 'y':
        print("Proceeding to create user accounts...")
        ac = AccessCreator(name, title)
        ac.createAll()
        ac.teardown()
        print("Complete!")


if __name__ == "__main__":
    main()
